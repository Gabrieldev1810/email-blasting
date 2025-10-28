import logging

logger = logging.getLogger(__name__)
from flask import Blueprint, jsonify, g
from app import db
from app.models.user import User, UserRole
from app.models.contact import Contact, ContactStatus
from app.models.campaign import Campaign, CampaignStatus
from app.models.email_log import EmailLog, EmailStatus, LinkClick
from app.middleware.auth import authenticated_required
from sqlalchemy import func

bp = Blueprint('dashboard', __name__)

@bp.route('/health', methods=['GET'])
def health():
    """Health check endpoint for dashboard routes."""
    return {'status': 'ok', 'service': 'dashboard'}

@bp.route('/stats', methods=['GET'])
@authenticated_required
def get_dashboard_stats():
    """Get comprehensive dashboard statistics."""
    try:
        # Get current authenticated user
        current_user = g.current_user
        
        # Admin users see all stats, regular users see only their own
        if current_user.role == UserRole.ADMIN:
            # Total users count
            total_users = User.query.count()
            
            # Total contacts count and status breakdown
            total_contacts = Contact.query.count()
            active_contacts = Contact.query.filter_by(status=ContactStatus.ACTIVE).count()
            unsubscribed_contacts = Contact.query.filter_by(status=ContactStatus.UNSUBSCRIBED).count()
            bounced_contacts = Contact.query.filter_by(status=ContactStatus.BOUNCED).count()
            
            # Campaign stats
            total_campaigns = Campaign.query.count()
            campaigns = Campaign.query.all()
            
            # Email stats from campaigns (legacy)
            total_emails_sent_legacy = sum(campaign.emails_sent or 0 for campaign in campaigns)
            total_emails_opened_legacy = sum(campaign.emails_opened or 0 for campaign in campaigns)
            total_emails_clicked_legacy = sum(campaign.emails_clicked or 0 for campaign in campaigns)
            
            # Email engagement stats from EmailLog (more accurate)
            total_emails_sent = EmailLog.query.count()
            total_emails_opened = EmailLog.query.filter(EmailLog.status.in_([EmailStatus.OPENED, EmailStatus.CLICKED])).count()
            total_emails_clicked = EmailLog.query.filter(EmailLog.status == EmailStatus.CLICKED).count()
            total_emails_bounced = EmailLog.query.filter(EmailLog.status == EmailStatus.BOUNCED).count()
            
            # Recent campaigns
            recent_campaigns = Campaign.query.order_by(Campaign.created_at.desc()).limit(5).all()
            
        else:
            # Regular users see only their own data
            total_users = 1  # Current user only
            
            # User's own contacts and status breakdown
            total_contacts = Contact.query.filter_by(user_id=current_user.id).count()
            active_contacts = Contact.query.filter_by(
                user_id=current_user.id, 
                status=ContactStatus.ACTIVE
            ).count()
            unsubscribed_contacts = Contact.query.filter_by(
                user_id=current_user.id, 
                status=ContactStatus.UNSUBSCRIBED
            ).count()
            bounced_contacts = Contact.query.filter_by(
                user_id=current_user.id, 
                status=ContactStatus.BOUNCED
            ).count()
            
            # User's campaigns
            total_campaigns = Campaign.query.filter_by(user_id=current_user.id).count()
            campaigns = Campaign.query.filter_by(user_id=current_user.id).all()
            
            # Email stats from user's campaigns (legacy)
            total_emails_sent_legacy = sum(campaign.emails_sent or 0 for campaign in campaigns)
            total_emails_opened_legacy = sum(campaign.emails_opened or 0 for campaign in campaigns)
            total_emails_clicked_legacy = sum(campaign.emails_clicked or 0 for campaign in campaigns)
            
            # Email engagement stats from EmailLog for this user (more accurate)
            user_email_logs = EmailLog.query.join(Campaign).filter(Campaign.user_id == current_user.id)
            total_emails_sent = user_email_logs.count()
            total_emails_opened = user_email_logs.filter(EmailLog.status.in_([EmailStatus.OPENED, EmailStatus.CLICKED])).count()
            total_emails_clicked = user_email_logs.filter(EmailLog.status == EmailStatus.CLICKED).count()
            total_emails_bounced = user_email_logs.filter(EmailLog.status == EmailStatus.BOUNCED).count()
            
            # Recent campaigns for this user
            recent_campaigns = Campaign.query.filter_by(user_id=current_user.id)\
                .order_by(Campaign.created_at.desc()).limit(5).all()
        
        # Format recent campaigns
        recent_campaigns_data = []
        for campaign in recent_campaigns:
            recent_campaigns_data.append({
                'id': campaign.id,
                'name': campaign.name,
                'status': campaign.status.value,
                'emails_sent': campaign.emails_sent or 0,
                'emails_opened': campaign.emails_opened or 0,
                'emails_clicked': campaign.emails_clicked or 0,
                'created_at': campaign.created_at.isoformat() if campaign.created_at else None
            })
        
        return jsonify({
            'success': True,
            'data': {
                # User and contact counts
                'total_users': total_users,
                'total_contacts': total_contacts,
                'active_contacts': active_contacts,
                'unsubscribed_contacts': unsubscribed_contacts,
                'bounced_contacts': bounced_contacts,
                
                # Campaign stats
                'total_campaigns': total_campaigns,
                'total_emails_sent': total_emails_sent,
                'total_emails_opened': total_emails_opened,
                'total_emails_clicked': total_emails_clicked,
                'total_emails_bounced': total_emails_bounced,
                
                # Engagement rates
                'email_open_rate': round((total_emails_opened / total_emails_sent) * 100, 2) if total_emails_sent > 0 else 0,
                'email_click_rate': round((total_emails_clicked / total_emails_sent) * 100, 2) if total_emails_sent > 0 else 0,
                'email_bounce_rate': round((total_emails_bounced / total_emails_sent) * 100, 2) if total_emails_sent > 0 else 0,
                
                # Recent activity
                'recent_campaigns': recent_campaigns_data
            }
        })
        
    except Exception as e:
        logger.error(f"Error fetching dashboard stats: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@bp.route('/lead-generation', methods=['GET'])
@authenticated_required
def get_lead_generation_data():
    """Get lead generation analytics data."""
    try:
        from datetime import datetime, timedelta
        from sqlalchemy import and_, extract
        
        current_user = g.current_user
        
        # Get data for the last 7 days
        end_date = datetime.now()
        start_date = end_date - timedelta(days=6)  # 7 days including today
        
        # Get campaign data for the last 7 days
        if current_user.role == UserRole.ADMIN:
            campaigns = Campaign.query.filter(
                Campaign.created_at >= start_date,
                Campaign.created_at <= end_date
            ).all()
        else:
            campaigns = Campaign.query.filter(
                and_(
                    Campaign.user_id == current_user.id,
                    Campaign.created_at >= start_date,
                    Campaign.created_at <= end_date
                )
            ).all()
        
        # Create data structure for each day
        lead_data = []
        for i in range(7):
            current_date = start_date + timedelta(days=i)
            day_name = current_date.strftime('%a')  # Mon, Tue, etc.
            
            # Count campaigns created on this day (representing email lead generation)
            email_campaigns = len([c for c in campaigns 
                                 if c.created_at.date() == current_date.date()])
            
            # Count contacts added on this day (representing organic leads)
            if current_user.role == UserRole.ADMIN:
                organic_leads = Contact.query.filter(
                    func.date(Contact.created_at) == current_date.date()
                ).count()
            else:
                organic_leads = Contact.query.filter(
                    and_(
                        Contact.user_id == current_user.id,
                        func.date(Contact.created_at) == current_date.date()
                    )
                ).count()
            
            lead_data.append({
                'day': day_name,
                'date': current_date.strftime('%Y-%m-%d'),
                'email_campaigns': email_campaigns * 10,  # Multiply for better visualization
                'organic_leads': organic_leads
            })
        
        # Calculate totals and conversion rates
        total_email_leads = sum(item['email_campaigns'] for item in lead_data)
        total_organic_leads = sum(item['organic_leads'] for item in lead_data)
        total_leads = total_email_leads + total_organic_leads
        
        # Calculate growth rate (compare with previous week)
        previous_week_start = start_date - timedelta(days=7)
        previous_week_end = start_date - timedelta(days=1)
        
        if current_user.role == UserRole.ADMIN:
            previous_week_contacts = Contact.query.filter(
                Contact.created_at >= previous_week_start,
                Contact.created_at <= previous_week_end
            ).count()
        else:
            previous_week_contacts = Contact.query.filter(
                and_(
                    Contact.user_id == current_user.id,
                    Contact.created_at >= previous_week_start,
                    Contact.created_at <= previous_week_end
                )
            ).count()
        
        current_week_contacts = total_organic_leads
        growth_rate = 0
        if previous_week_contacts > 0:
            growth_rate = ((current_week_contacts - previous_week_contacts) / previous_week_contacts) * 100
        
        return jsonify({
            'success': True,
            'data': {
                'weekly_data': lead_data,
                'totals': {
                    'total_leads': total_leads,
                    'email_leads': total_email_leads,
                    'organic_leads': total_organic_leads
                },
                'metrics': {
                    'conversion_rate': round((total_organic_leads / max(total_leads, 1)) * 100, 1),
                    'growth_rate': round(growth_rate, 1),
                    'avg_daily_leads': round(total_leads / 7, 1)
                }
            }
        })
        
    except Exception as e:
        logger.error(f"Error fetching lead generation data: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@bp.route('/email-analytics', methods=['GET'])
@authenticated_required
def get_email_analytics():
    """Get detailed email analytics data."""
    try:
        from datetime import datetime, timedelta
        from sqlalchemy import and_, func
        
        current_user = g.current_user
        
        # Get data for the last 30 days
        end_date = datetime.now()
        start_date = end_date - timedelta(days=29)  # 30 days including today
        
        # Get campaigns data for analytics
        if current_user.role == UserRole.ADMIN:
            campaigns = Campaign.query.filter(
                Campaign.created_at >= start_date,
                Campaign.created_at <= end_date
            ).all()
            
            # Get all contacts for unsubscribe analytics
            total_contacts = Contact.query.count()
            unsubscribed_contacts = Contact.query.filter_by(status=ContactStatus.UNSUBSCRIBED).count()
            
        else:
            campaigns = Campaign.query.filter(
                and_(
                    Campaign.user_id == current_user.id,
                    Campaign.created_at >= start_date,
                    Campaign.created_at <= end_date
                )
            ).all()
            
            # Get user's contacts for unsubscribe analytics
            total_contacts = Contact.query.filter_by(user_id=current_user.id).count()
            unsubscribed_contacts = Contact.query.filter_by(
                user_id=current_user.id, 
                status=ContactStatus.UNSUBSCRIBED
            ).count()
        
        # Create daily analytics data
        daily_data = []
        total_opens = 0
        total_clicks = 0
        total_sent = 0
        
        for i in range(30):
            current_date = start_date + timedelta(days=i)
            day_campaigns = [c for c in campaigns if c.created_at.date() == current_date.date()]
            
            daily_opens = sum(c.emails_opened or 0 for c in day_campaigns)
            daily_clicks = sum(c.emails_clicked or 0 for c in day_campaigns)
            daily_sent = sum(c.emails_sent or 0 for c in day_campaigns)
            
            total_opens += daily_opens
            total_clicks += daily_clicks
            total_sent += daily_sent
            
            daily_data.append({
                'date': current_date.strftime('%Y-%m-%d'),
                'day': current_date.strftime('%a'),
                'opens': daily_opens,
                'clicks': daily_clicks,
                'sent': daily_sent,
                'open_rate': round((daily_opens / max(daily_sent, 1)) * 100, 1),
                'click_rate': round((daily_clicks / max(daily_sent, 1)) * 100, 1)
            })
        
        # Calculate metrics
        overall_open_rate = round((total_opens / max(total_sent, 1)) * 100, 1)
        overall_click_rate = round((total_clicks / max(total_sent, 1)) * 100, 1)
        unsubscribe_rate = round((unsubscribed_contacts / max(total_contacts, 1)) * 100, 1)
        
        # Calculate trends (compare with previous 30 days)
        previous_start = start_date - timedelta(days=30)
        previous_end = start_date - timedelta(days=1)
        
        if current_user.role == UserRole.ADMIN:
            previous_campaigns = Campaign.query.filter(
                Campaign.created_at >= previous_start,
                Campaign.created_at <= previous_end
            ).all()
        else:
            previous_campaigns = Campaign.query.filter(
                and_(
                    Campaign.user_id == current_user.id,
                    Campaign.created_at >= previous_start,
                    Campaign.created_at <= previous_end
                )
            ).all()
        
        previous_opens = sum(c.emails_opened or 0 for c in previous_campaigns)
        previous_clicks = sum(c.emails_clicked or 0 for c in previous_campaigns)
        
        opens_trend = 0 if previous_opens == 0 else round(((total_opens - previous_opens) / previous_opens) * 100, 1)
        clicks_trend = 0 if previous_clicks == 0 else round(((total_clicks - previous_clicks) / previous_clicks) * 100, 1)
        
        return jsonify({
            'success': True,
            'data': {
                'daily_analytics': daily_data,
                'totals': {
                    'total_opens': total_opens,
                    'total_clicks': total_clicks,
                    'total_sent': total_sent,
                    'unsubscribed': unsubscribed_contacts
                },
                'metrics': {
                    'open_rate': overall_open_rate,
                    'click_rate': overall_click_rate,
                    'unsubscribe_rate': unsubscribe_rate,
                    'opens_trend': opens_trend,
                    'clicks_trend': clicks_trend
                }
            }
        })
        
    except Exception as e:
        logger.error(f"Error fetching email analytics data: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@bp.route('/email-engagement', methods=['GET'])
@authenticated_required
def get_email_engagement_metrics():
    """Get comprehensive email engagement metrics from EmailLog data."""
    try:
        from datetime import datetime, timedelta
        from sqlalchemy import and_
        
        current_user = g.current_user
        
        # Build base query based on user role
        if current_user.role == UserRole.ADMIN:
            # Admin sees all data
            base_query = EmailLog.query.join(Campaign)
        else:
            # Regular users see only their campaign data
            base_query = EmailLog.query.join(Campaign).filter(Campaign.user_id == current_user.id)
        
        # Overall engagement metrics
        total_sent = base_query.count()
        total_opened = base_query.filter(EmailLog.status.in_([EmailStatus.OPENED, EmailStatus.CLICKED])).count()
        total_clicked = base_query.filter(EmailLog.status == EmailStatus.CLICKED).count()
        total_bounced = base_query.filter(EmailLog.status == EmailStatus.BOUNCED).count()
        
        # Calculate rates
        open_rate = round((total_opened / total_sent) * 100, 2) if total_sent > 0 else 0
        click_rate = round((total_clicked / total_sent) * 100, 2) if total_sent > 0 else 0
        bounce_rate = round((total_bounced / total_sent) * 100, 2) if total_sent > 0 else 0
        
        # Get data for the last 30 days for trends
        thirty_days_ago = datetime.utcnow() - timedelta(days=30)
        
        # Daily engagement data for charts
        daily_engagement = []
        for i in range(7):  # Last 7 days
            date = datetime.utcnow() - timedelta(days=6-i)
            day_start = date.replace(hour=0, minute=0, second=0, microsecond=0)
            day_end = date.replace(hour=23, minute=59, second=59, microsecond=999999)
            
            day_sent = base_query.filter(
                EmailLog.sent_at >= day_start,
                EmailLog.sent_at <= day_end
            ).count()
            
            day_opened = base_query.filter(
                EmailLog.opened_at >= day_start,
                EmailLog.opened_at <= day_end
            ).count()
            
            day_clicked = base_query.filter(
                EmailLog.clicked_at >= day_start,
                EmailLog.clicked_at <= day_end
            ).count()
            
            day_bounced = base_query.filter(
                EmailLog.bounced_at >= day_start,
                EmailLog.bounced_at <= day_end
            ).count()
            
            daily_engagement.append({
                'date': date.strftime('%Y-%m-%d'),
                'day': date.strftime('%a'),
                'sent': day_sent,
                'opened': day_opened,
                'clicked': day_clicked,
                'bounced': day_bounced
            })
        
        # Top performing campaigns by engagement
        if current_user.role == UserRole.ADMIN:
            campaigns_query = Campaign.query
        else:
            campaigns_query = Campaign.query.filter_by(user_id=current_user.id)
        
        top_campaigns = []
        for campaign in campaigns_query.order_by(Campaign.created_at.desc()).limit(10):
            campaign_logs = EmailLog.query.filter_by(campaign_id=campaign.id)
            campaign_sent = campaign_logs.count()
            campaign_opened = campaign_logs.filter(EmailLog.status.in_([EmailStatus.OPENED, EmailStatus.CLICKED])).count()
            campaign_clicked = campaign_logs.filter(EmailLog.status == EmailStatus.CLICKED).count()
            campaign_bounced = campaign_logs.filter(EmailLog.status == EmailStatus.BOUNCED).count()
            
            if campaign_sent > 0:
                campaign_open_rate = round((campaign_opened / campaign_sent) * 100, 2)
                campaign_click_rate = round((campaign_clicked / campaign_sent) * 100, 2)
                campaign_bounce_rate = round((campaign_bounced / campaign_sent) * 100, 2)
            else:
                campaign_open_rate = campaign_click_rate = campaign_bounce_rate = 0
            
            top_campaigns.append({
                'id': campaign.id,
                'name': campaign.name,
                'sent': campaign_sent,
                'opened': campaign_opened,
                'clicked': campaign_clicked,
                'bounced': campaign_bounced,
                'open_rate': campaign_open_rate,
                'click_rate': campaign_click_rate,
                'bounce_rate': campaign_bounce_rate,
                'created_at': campaign.created_at.isoformat() if campaign.created_at else None
            })
        
        return jsonify({
            'success': True,
            'data': {
                'overview': {
                    'total_sent': total_sent,
                    'total_opened': total_opened,
                    'total_clicked': total_clicked,
                    'total_bounced': total_bounced,
                    'open_rate': open_rate,
                    'click_rate': click_rate,
                    'bounce_rate': bounce_rate
                },
                'daily_engagement': daily_engagement,
                'top_campaigns': top_campaigns
            }
        })
        
    except Exception as e:
        logger.error(f"Error fetching email engagement metrics: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@bp.route('/campaign/<int:campaign_id>/recipients', methods=['GET'])
@authenticated_required
def get_campaign_recipients(campaign_id):
    """Get detailed recipient tracking data for a specific campaign."""
    try:
        from flask import request
        
        current_user = g.current_user
        
        # Verify campaign belongs to user (unless admin)
        campaign = Campaign.query.get_or_404(campaign_id)
        if current_user.role != UserRole.ADMIN and campaign.user_id != current_user.id:
            return jsonify({'success': False, 'error': 'Unauthorized'}), 403
        
        # Get filter parameters
        status_filter = request.args.get('status')  # sent, opened, clicked, bounced
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 50))
        
        # Build query
        query = EmailLog.query.filter_by(campaign_id=campaign_id)
        
        if status_filter:
            if status_filter == 'opened':
                query = query.filter(EmailLog.status.in_([EmailStatus.OPENED, EmailStatus.CLICKED]))
            elif status_filter == 'clicked':
                query = query.filter(EmailLog.status == EmailStatus.CLICKED)
            elif status_filter == 'bounced':
                query = query.filter(EmailLog.status == EmailStatus.BOUNCED)
            elif status_filter == 'sent':
                query = query.filter(EmailLog.status == EmailStatus.SENT)
        
        # Paginate results
        paginated = query.paginate(page=page, per_page=per_page, error_out=False)
        
        # Format recipient data
        recipients = []
        for log in paginated.items:
            # Get click count for this recipient
            click_count = LinkClick.query.filter_by(email_log_id=log.id).count()
            
            recipients.append({
                'id': log.id,
                'email': log.recipient_email,
                'name': log.recipient_name,
                'status': log.status.value,
                'sent_at': log.sent_at.isoformat() if log.sent_at else None,
                'opened_at': log.opened_at.isoformat() if log.opened_at else None,
                'clicked_at': log.clicked_at.isoformat() if log.clicked_at else None,
                'bounced_at': log.bounced_at.isoformat() if log.bounced_at else None,
                'bounce_type': log.bounce_type.value if log.bounce_type else None,
                'bounce_reason': log.bounce_reason,
                'click_count': click_count,
                'tracking_id': log.tracking_id
            })
        
        return jsonify({
            'success': True,
            'data': {
                'recipients': recipients,
                'pagination': {
                    'page': page,
                    'per_page': per_page,
                    'total': paginated.total,
                    'pages': paginated.pages,
                    'has_next': paginated.has_next,
                    'has_prev': paginated.has_prev
                }
            }
        })
        
    except Exception as e:
        logger.error(f"Error fetching campaign recipients: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@bp.route('/campaign/<int:campaign_id>/export', methods=['GET'])
@authenticated_required
def export_campaign_metrics(campaign_id):
    """Export campaign metrics as CSV."""
    try:
        from flask import Response
        import csv
        from io import StringIO
        
        current_user = g.current_user
        
        # Verify campaign belongs to user (unless admin)
        campaign = Campaign.query.get_or_404(campaign_id)
        if current_user.role != UserRole.ADMIN and campaign.user_id != current_user.id:
            return jsonify({'success': False, 'error': 'Unauthorized'}), 403
        
        # Get all email logs for this campaign
        email_logs = EmailLog.query.filter_by(campaign_id=campaign_id).all()
        
        # Create CSV content
        output = StringIO()
        writer = csv.writer(output)
        
        # Write headers
        writer.writerow([
            'Email', 'Name', 'Status', 'Sent At', 'Opened At', 'Clicked At', 
            'Bounced At', 'Bounce Type', 'Bounce Reason', 'Click Count'
        ])
        
        # Write data
        for log in email_logs:
            click_count = LinkClick.query.filter_by(email_log_id=log.id).count()
            writer.writerow([
                log.recipient_email,
                log.recipient_name or '',
                log.status.value,
                log.sent_at.isoformat() if log.sent_at else '',
                log.opened_at.isoformat() if log.opened_at else '',
                log.clicked_at.isoformat() if log.clicked_at else '',
                log.bounced_at.isoformat() if log.bounced_at else '',
                log.bounce_type.value if log.bounce_type else '',
                log.bounce_reason or '',
                click_count
            ])
        
        # Create response
        csv_content = output.getvalue()
        output.close()
        
        return Response(
            csv_content,
            mimetype='text/csv',
            headers={
                'Content-Disposition': f'attachment; filename=campaign_{campaign_id}_metrics.csv'
            }
        )
        
    except Exception as e:
        logger.error(f"Error exporting campaign metrics: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@bp.route('/email-engagement', methods=['GET'])
@authenticated_required
def get_email_engagement_data():
    """Get comprehensive email engagement metrics."""
    try:
        from datetime import datetime, timedelta
        from sqlalchemy import and_
        
        current_user = g.current_user
        
        # Base query for email logs
        if current_user.role == UserRole.ADMIN:
            base_query = EmailLog.query
        else:
            base_query = EmailLog.query.join(Campaign).filter(Campaign.user_id == current_user.id)
        
        # Calculate overall metrics
        total_sent = base_query.count()
        total_opened = base_query.filter(EmailLog.status.in_([EmailStatus.OPENED, EmailStatus.CLICKED])).count()
        total_clicked = base_query.filter(EmailLog.status == EmailStatus.CLICKED).count()
        total_bounced = base_query.filter(EmailLog.status == EmailStatus.BOUNCED).count()
        
        # Calculate rates
        open_rate = round((total_opened / total_sent) * 100, 2) if total_sent > 0 else 0
        click_rate = round((total_clicked / total_sent) * 100, 2) if total_sent > 0 else 0
        bounce_rate = round((total_bounced / total_sent) * 100, 2) if total_sent > 0 else 0
        
        # Get daily engagement data for the last 7 days
        end_date = datetime.now()
        start_date = end_date - timedelta(days=6)
        
        daily_data = []
        for i in range(7):
            current_date = start_date + timedelta(days=i)
            day_name = current_date.strftime('%a')
            
            # Get email logs for this day
            day_query = base_query.filter(
                func.date(EmailLog.sent_at) == current_date.date()
            )
            
            day_sent = day_query.count()
            day_opened = day_query.filter(EmailLog.status.in_([EmailStatus.OPENED, EmailStatus.CLICKED])).count()
            day_clicked = day_query.filter(EmailLog.status == EmailStatus.CLICKED).count()
            day_bounced = day_query.filter(EmailLog.status == EmailStatus.BOUNCED).count()
            
            daily_data.append({
                'day': day_name,
                'date': current_date.strftime('%Y-%m-%d'),
                'sent': day_sent,
                'opened': day_opened,
                'clicked': day_clicked,
                'bounced': day_bounced,
                'open_rate': round((day_opened / day_sent) * 100, 2) if day_sent > 0 else 0,
                'click_rate': round((day_clicked / day_sent) * 100, 2) if day_sent > 0 else 0,
                'bounce_rate': round((day_bounced / day_sent) * 100, 2) if day_sent > 0 else 0
            })
        
        return jsonify({
            'success': True,
            'data': {
                'overview': {
                    'total_sent': total_sent,
                    'total_opened': total_opened,
                    'total_clicked': total_clicked,
                    'total_bounced': total_bounced,
                    'open_rate': open_rate,
                    'click_rate': click_rate,
                    'bounce_rate': bounce_rate
                },
                'daily_engagement': daily_data,
                'trends': {
                    'total_engagement': total_opened + total_clicked,
                    'engagement_rate': round(((total_opened + total_clicked) / total_sent) * 100, 2) if total_sent > 0 else 0
                }
            }
        })
        
    except Exception as e:
        logger.error(f"Error fetching email engagement data: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@bp.route('/email-logs', methods=['GET'])
@authenticated_required
def get_email_logs():
    """Get paginated email logs with detailed recipient tracking."""
    try:
        from flask import request
        from datetime import datetime
        from sqlalchemy import and_, or_
        
        current_user = g.current_user
        
        # Get query parameters
        page = int(request.args.get('page', 1))
        per_page = min(int(request.args.get('per_page', 20)), 100)  # Max 100 per page
        campaign_id = request.args.get('campaign_id')
        status_filter = request.args.get('status')
        search_email = request.args.get('search_email')
        date_from = request.args.get('date_from')
        date_to = request.args.get('date_to')
        
        # Base query for email logs
        if current_user.role == UserRole.ADMIN:
            query = EmailLog.query.join(Campaign, EmailLog.campaign_id == Campaign.id)
        else:
            query = EmailLog.query.join(Campaign, EmailLog.campaign_id == Campaign.id).filter(Campaign.user_id == current_user.id)
        
        # Apply filters
        if campaign_id and campaign_id != 'all':
            query = query.filter(EmailLog.campaign_id == int(campaign_id))
        
        if status_filter and status_filter != 'all':
            query = query.filter(EmailLog.status == EmailStatus(status_filter))
        
        if search_email:
            query = query.filter(EmailLog.recipient_email.ilike(f'%{search_email}%'))
        
        if date_from:
            date_from_obj = datetime.fromisoformat(date_from.replace('Z', '+00:00'))
            query = query.filter(EmailLog.sent_at >= date_from_obj)
        
        if date_to:
            date_to_obj = datetime.fromisoformat(date_to.replace('Z', '+00:00'))
            query = query.filter(EmailLog.sent_at <= date_to_obj)
        
        # Get total count
        total_count = query.count()
        
        # Get paginated results
        email_logs = query.order_by(EmailLog.sent_at.desc()).offset((page - 1) * per_page).limit(per_page).all()
        
        # Format results
        logs_data = []
        for log in email_logs:
            # Get click count for this email log
            click_count = LinkClick.query.filter_by(email_log_id=log.id).count()
            
            logs_data.append({
                'id': log.id,
                'campaign_id': log.campaign_id,
                'campaign_name': log.campaign.name if log.campaign else 'Unknown',
                'recipient_email': log.recipient_email,
                'recipient_name': log.recipient_name,
                'status': log.status.value,
                'sent_at': log.sent_at.isoformat() if log.sent_at else None,
                'opened_at': log.opened_at.isoformat() if log.opened_at else None,
                'clicked_at': log.clicked_at.isoformat() if log.clicked_at else None,
                'bounced_at': log.bounced_at.isoformat() if log.bounced_at else None,
                'bounce_type': log.bounce_type.value if log.bounce_type else None,
                'bounce_reason': log.bounce_reason,
                'click_count': click_count,
                'tracking_id': log.tracking_id
            })
        
        return jsonify({
            'success': True,
            'data': {
                'logs': logs_data,
                'total': total_count,
                'page': page,
                'per_page': per_page,
                'total_pages': (total_count + per_page - 1) // per_page
            }
        })
        
    except Exception as e:
        logger.error(f"Error fetching email logs: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@bp.route('/email-logs/export', methods=['GET'])
@authenticated_required
def export_email_logs():
    """Export email logs to CSV."""
    try:
        from flask import request, Response
        from datetime import datetime
        from sqlalchemy import and_, or_
        import csv
        import io
        
        current_user = g.current_user
        
        # Get query parameters (same as get_email_logs)
        campaign_id = request.args.get('campaign_id')
        status_filter = request.args.get('status')
        search_email = request.args.get('search_email')
        date_from = request.args.get('date_from')
        date_to = request.args.get('date_to')
        
        # Base query for email logs
        if current_user.role == UserRole.ADMIN:
            query = EmailLog.query.join(Campaign, EmailLog.campaign_id == Campaign.id)
        else:
            query = EmailLog.query.join(Campaign, EmailLog.campaign_id == Campaign.id).filter(Campaign.user_id == current_user.id)
        
        # Apply same filters as get_email_logs
        if campaign_id and campaign_id != 'all':
            query = query.filter(EmailLog.campaign_id == int(campaign_id))
        
        if status_filter and status_filter != 'all':
            query = query.filter(EmailLog.status == EmailStatus(status_filter))
        
        if search_email:
            query = query.filter(EmailLog.recipient_email.ilike(f'%{search_email}%'))
        
        if date_from:
            date_from_obj = datetime.fromisoformat(date_from.replace('Z', '+00:00'))
            query = query.filter(EmailLog.sent_at >= date_from_obj)
        
        if date_to:
            date_to_obj = datetime.fromisoformat(date_to.replace('Z', '+00:00'))
            query = query.filter(EmailLog.sent_at <= date_to_obj)
        
        # Get all results (limit to 10000 for performance)
        email_logs = query.order_by(EmailLog.sent_at.desc()).limit(10000).all()
        
        # Create CSV
        output = io.StringIO()
        fieldnames = [
            'Campaign Name', 'Recipient Email', 'Recipient Name', 'Status',
            'Sent At', 'Opened At', 'Clicked At', 'Bounced At',
            'Bounce Type', 'Bounce Reason', 'Click Count', 'Tracking ID'
        ]
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        
        for log in email_logs:
            click_count = LinkClick.query.filter_by(email_log_id=log.id).count()
            
            writer.writerow({
                'Campaign Name': log.campaign.name if log.campaign else 'Unknown',
                'Recipient Email': log.recipient_email,
                'Recipient Name': log.recipient_name or '',
                'Status': log.status.value,
                'Sent At': log.sent_at.isoformat() if log.sent_at else '',
                'Opened At': log.opened_at.isoformat() if log.opened_at else '',
                'Clicked At': log.clicked_at.isoformat() if log.clicked_at else '',
                'Bounced At': log.bounced_at.isoformat() if log.bounced_at else '',
                'Bounce Type': log.bounce_type.value if log.bounce_type else '',
                'Bounce Reason': log.bounce_reason or '',
                'Click Count': click_count,
                'Tracking ID': log.tracking_id
            })
        
        # Create response
        csv_content = output.getvalue()
        output.close()
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'email_recipients_{campaign_id or "all"}_{timestamp}.csv'
        
        return Response(
            csv_content,
            mimetype='text/csv',
            headers={
                'Content-Disposition': f'attachment; filename={filename}'
            }
        )
        
    except Exception as e:
        logger.error(f"Error exporting email logs: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@bp.route('/export-analytics-test', methods=['GET', 'POST'])
def test_export_endpoint():
    """Test endpoint to verify route registration."""
    from flask import request
    return jsonify({'success': True, 'message': 'Export endpoint is working', 'method': request.method})

@bp.route('/export-analytics', methods=['POST'])
@authenticated_required
def export_dashboard_analytics():
    """Export comprehensive dashboard analytics with filtering options."""
    try:
        from flask import request, Response
        from datetime import datetime, timedelta
        import json
        import csv
        import io
        
        logger.info(f"Export analytics endpoint called")
        current_user = g.current_user
        logger.debug(f"Current user: {current_user.email}")
        
        # First, let's add some basic debugging
        try:
            data = request.get_json()
            logger.debug(f"Received data: {data}")
        except Exception as e:
            logger.error(f"Error getting JSON data: {e}")
            return jsonify({'success': False, 'error': f'JSON parsing error: {str(e)}'}), 400
        
        if not data:
            return jsonify({'success': False, 'error': 'No data provided'}), 400
        
        # Parse filters
        date_from = data.get('dateFrom')
        date_to = data.get('dateTo')
        report_types = data.get('reportTypes', [])
        format_type = data.get('format', 'csv')  # csv or excel
        
        logger.debug(f"Parsed filters - dateFrom: {date_from}, dateTo: {date_to}, reportTypes: {report_types}, format: {format_type}")
        
        # Set default date range if not provided (last 30 days)
        if not date_from or not date_to:
            date_to = datetime.utcnow()
            date_from = date_to - timedelta(days=30)
        else:
            date_from = datetime.fromisoformat(date_from.replace('Z', '+00:00'))
            date_to = datetime.fromisoformat(date_to.replace('Z', '+00:00'))
        
        # Prepare data collection based on selected report types
        export_data = {}
        
        # Overview Stats
        if 'overview' in report_types:
            if current_user.role == UserRole.ADMIN:
                total_contacts = Contact.query.count()
                total_campaigns = Campaign.query.count()
            else:
                total_contacts = Contact.query.filter_by(user_id=current_user.id).count()
                total_campaigns = Campaign.query.filter_by(user_id=current_user.id).count()
            
            # Email stats within date range
            email_query = EmailLog.query.filter(
                EmailLog.sent_at >= date_from,
                EmailLog.sent_at <= date_to
            )
            
            if current_user.role != UserRole.ADMIN:
                email_query = email_query.join(Campaign).filter(Campaign.user_id == current_user.id)
            
            total_emails_sent = email_query.count()
            total_emails_opened = email_query.filter(EmailLog.status.in_([EmailStatus.OPENED, EmailStatus.CLICKED])).count()
            total_emails_clicked = email_query.filter(EmailLog.status == EmailStatus.CLICKED).count()
            total_emails_bounced = email_query.filter(EmailLog.status == EmailStatus.BOUNCED).count()
            
            # Calculate rates
            open_rate = (total_emails_opened / total_emails_sent * 100) if total_emails_sent > 0 else 0
            click_rate = (total_emails_clicked / total_emails_sent * 100) if total_emails_sent > 0 else 0
            bounce_rate = (total_emails_bounced / total_emails_sent * 100) if total_emails_sent > 0 else 0
            
            export_data['overview'] = {
                'Total Contacts': total_contacts,
                'Total Campaigns': total_campaigns,
                'Emails Sent': total_emails_sent,
                'Emails Opened': total_emails_opened,
                'Emails Clicked': total_emails_clicked,
                'Emails Bounced': total_emails_bounced,
                'Open Rate (%)': round(open_rate, 2),
                'Click Rate (%)': round(click_rate, 2),
                'Bounce Rate (%)': round(bounce_rate, 2)
            }
        
        # Campaign Performance
        if 'campaigns' in report_types:
            campaigns_query = Campaign.query
            if current_user.role != UserRole.ADMIN:
                campaigns_query = campaigns_query.filter_by(user_id=current_user.id)
            
            campaigns = campaigns_query.all()
            campaign_data = []
            
            for campaign in campaigns:
                # Get email logs for this campaign within date range
                email_logs = EmailLog.query.filter(
                    EmailLog.campaign_id == campaign.id,
                    EmailLog.sent_at >= date_from,
                    EmailLog.sent_at <= date_to
                ).all()
                
                sent = len(email_logs)
                opened = len([log for log in email_logs if log.status in [EmailStatus.OPENED, EmailStatus.CLICKED]])
                clicked = len([log for log in email_logs if log.status == EmailStatus.CLICKED])
                bounced = len([log for log in email_logs if log.status == EmailStatus.BOUNCED])
                
                open_rate = (opened / sent * 100) if sent > 0 else 0
                click_rate = (clicked / sent * 100) if sent > 0 else 0
                bounce_rate = (bounced / sent * 100) if sent > 0 else 0
                
                campaign_data.append({
                    'Campaign Name': campaign.name,
                    'Status': campaign.status.value,
                    'Created': campaign.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                    'Emails Sent': sent,
                    'Emails Opened': opened,
                    'Emails Clicked': clicked,
                    'Emails Bounced': bounced,
                    'Open Rate (%)': round(open_rate, 2),
                    'Click Rate (%)': round(click_rate, 2),
                    'Bounce Rate (%)': round(bounce_rate, 2)
                })
            
            export_data['campaigns'] = campaign_data
        
        # Email Engagement
        if 'engagement' in report_types:
            # Get email engagement data grouped by date
            email_query = EmailLog.query.filter(
                EmailLog.sent_at >= date_from,
                EmailLog.sent_at <= date_to
            )
            
            if current_user.role != UserRole.ADMIN:
                email_query = email_query.join(Campaign).filter(Campaign.user_id == current_user.id)
            
            email_logs = email_query.all()
            
            # Group by date
            engagement_by_date = {}
            for log in email_logs:
                date_key = log.sent_at.strftime('%Y-%m-%d')
                if date_key not in engagement_by_date:
                    engagement_by_date[date_key] = {
                        'sent': 0, 'opened': 0, 'clicked': 0, 'bounced': 0
                    }
                
                engagement_by_date[date_key]['sent'] += 1
                if log.status in [EmailStatus.OPENED, EmailStatus.CLICKED]:
                    engagement_by_date[date_key]['opened'] += 1
                if log.status == EmailStatus.CLICKED:
                    engagement_by_date[date_key]['clicked'] += 1
                if log.status == EmailStatus.BOUNCED:
                    engagement_by_date[date_key]['bounced'] += 1
            
            engagement_data = []
            for date_str, stats in sorted(engagement_by_date.items()):
                open_rate = (stats['opened'] / stats['sent'] * 100) if stats['sent'] > 0 else 0
                click_rate = (stats['clicked'] / stats['sent'] * 100) if stats['sent'] > 0 else 0
                bounce_rate = (stats['bounced'] / stats['sent'] * 100) if stats['sent'] > 0 else 0
                
                engagement_data.append({
                    'Date': date_str,
                    'Emails Sent': stats['sent'],
                    'Emails Opened': stats['opened'],
                    'Emails Clicked': stats['clicked'],
                    'Emails Bounced': stats['bounced'],
                    'Open Rate (%)': round(open_rate, 2),
                    'Click Rate (%)': round(click_rate, 2),
                    'Bounce Rate (%)': round(bounce_rate, 2)
                })
            
            export_data['engagement'] = engagement_data
        
        # Contact Analytics
        if 'contacts' in report_types:
            contacts_query = Contact.query
            if current_user.role != UserRole.ADMIN:
                contacts_query = contacts_query.filter_by(user_id=current_user.id)
            
            contacts = contacts_query.all()
            contact_data = []
            
            for contact in contacts:
                # Get email statistics for this contact
                contact_emails = EmailLog.query.filter(
                    EmailLog.recipient_email == contact.email,
                    EmailLog.sent_at >= date_from,
                    EmailLog.sent_at <= date_to
                )
                
                if current_user.role != UserRole.ADMIN:
                    contact_emails = contact_emails.join(Campaign).filter(Campaign.user_id == current_user.id)
                
                contact_logs = contact_emails.all()
                
                sent = len(contact_logs)
                opened = len([log for log in contact_logs if log.status in [EmailStatus.OPENED, EmailStatus.CLICKED]])
                clicked = len([log for log in contact_logs if log.status == EmailStatus.CLICKED])
                bounced = len([log for log in contact_logs if log.status == EmailStatus.BOUNCED])
                
                # Combine first_name and last_name to create full name
                full_name = ''
                if contact.first_name and contact.last_name:
                    full_name = f"{contact.first_name} {contact.last_name}"
                elif contact.first_name:
                    full_name = contact.first_name
                elif contact.last_name:
                    full_name = contact.last_name
                
                contact_data.append({
                    'Email': contact.email,
                    'Name': full_name,
                    'Status': contact.status.value,
                    'Created': contact.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                    'Emails Received': sent,
                    'Emails Opened': opened,
                    'Emails Clicked': clicked,
                    'Emails Bounced': bounced
                })
            
            export_data['contacts'] = contact_data
        
        # Lead Generation
        if 'leads' in report_types:
            # Get lead generation data (contacts created within date range)
            leads_query = Contact.query.filter(
                Contact.created_at >= date_from,
                Contact.created_at <= date_to
            )
            
            if current_user.role != UserRole.ADMIN:
                leads_query = leads_query.filter_by(user_id=current_user.id)
            
            leads = leads_query.all()
            
            # Group by date
            leads_by_date = {}
            for contact in leads:
                date_key = contact.created_at.strftime('%Y-%m-%d')
                if date_key not in leads_by_date:
                    leads_by_date[date_key] = 0
                leads_by_date[date_key] += 1
            
            leads_data = []
            for date_str, count in sorted(leads_by_date.items()):
                leads_data.append({
                    'Date': date_str,
                    'New Contacts': count
                })
            
            export_data['leads'] = leads_data
        
        # Generate export file
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        if format_type == 'excel':
            # Create Excel file
            output = io.BytesIO()
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            
            workbook = Workbook()
            # Remove default sheet
            workbook.remove(workbook.active)
            
            # Create worksheets for each report type
            for report_type, data in export_data.items():
                if isinstance(data, list) and data:
                    # Create worksheet
                    worksheet = workbook.create_sheet(title=report_type.title())
                    
                    # Write headers with styling
                    headers = list(data[0].keys())
                    for col, header in enumerate(headers, start=1):
                        cell = worksheet.cell(row=1, column=col, value=header)
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="D7E4BC", end_color="D7E4BC", fill_type="solid")
                    
                    # Write data
                    for row, record in enumerate(data, start=2):
                        for col, header in enumerate(headers, start=1):
                            worksheet.cell(row=row, column=col, value=record[header])
                    
                    # Auto-adjust column widths
                    for col in range(1, len(headers) + 1):
                        worksheet.column_dimensions[chr(64 + col)].width = 15
                
                elif isinstance(data, dict):
                    # Create worksheet for overview data
                    worksheet = workbook.create_sheet(title=report_type.title())
                    
                    # Write overview data
                    row = 1
                    for key, value in data.items():
                        key_cell = worksheet.cell(row=row, column=1, value=key)
                        key_cell.font = Font(bold=True)
                        key_cell.fill = PatternFill(start_color="D7E4BC", end_color="D7E4BC", fill_type="solid")
                        worksheet.cell(row=row, column=2, value=value)
                        row += 1
                    
                    worksheet.column_dimensions['A'].width = 25
                    worksheet.column_dimensions['B'].width = 15
            
            workbook.save(output)
            excel_content = output.getvalue()
            output.close()
            
            filename = f'dashboard_analytics_{timestamp}.xlsx'
            
            return Response(
                excel_content,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                headers={
                    'Content-Disposition': f'attachment; filename={filename}'
                }
            )
        
        else:
            # Create CSV file (all data in one file)
            output = io.StringIO()
            
            # Write each report type as a section
            for report_type, data in export_data.items():
                output.write(f"\n# {report_type.upper()} REPORT\n")
                
                if isinstance(data, list) and data:
                    # Write CSV for list data
                    fieldnames = list(data[0].keys())
                    writer = csv.DictWriter(output, fieldnames=fieldnames)
                    writer.writeheader()
                    for record in data:
                        writer.writerow(record)
                
                elif isinstance(data, dict):
                    # Write overview data
                    for key, value in data.items():
                        output.write(f"{key},{value}\n")
                
                output.write("\n")
            
            csv_content = output.getvalue()
            output.close()
            
            filename = f'dashboard_analytics_{timestamp}.csv'
            
            return Response(
                csv_content,
                mimetype='text/csv',
                headers={
                    'Content-Disposition': f'attachment; filename={filename}'
                }
            )
    
    except Exception as e:
        logger.error(f"Error exporting dashboard analytics: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500